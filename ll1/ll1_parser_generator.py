"""
LL(1) Predictive Parser Generator for RecipeScript
Computes FIRST, FOLLOW, NULLABLE sets and generates Predictive Parsing Table
Outputs results to Excel file
"""

import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class LL1ParserGenerator:
    def __init__(self):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.nullable = {}
        self.first = defaultdict(set)
        self.follow = defaultdict(set)
        self.parsing_table = {}
        self.epsilon = 'ε'
        self.end_marker = '$'
        
    def add_production(self, lhs, rhs_list):
        """Add a production rule: lhs -> rhs1 | rhs2 | ..."""
        if lhs not in self.grammar:
            self.grammar[lhs] = []
        self.non_terminals.add(lhs)
        
        for rhs in rhs_list:
            # rhs is a list of symbols
            self.grammar[lhs].append(rhs)
            for symbol in rhs:
                if symbol != self.epsilon and not self.is_terminal(symbol):
                    self.non_terminals.add(symbol)
    
    def is_terminal(self, symbol):
        """Check if symbol is terminal"""
        if symbol == self.epsilon or symbol == self.end_marker:
            return False
        # Terminals are lowercase keywords, operators, or uppercase tokens
        return (symbol.isupper() or 
                symbol in ['+', '-', '*', '/', '=', '==', '!=', '>', '<', '>=', '<=',
                          ';', ',', '(', ')', '{', '}', '"'] or
                symbol.startswith('"') or
                symbol in ['ingredient', 'time', 'temp', 'quantity', 'text',
                          'mix', 'heat', 'wait', 'serve', 'display', 'add', 'scale',
                          'repeat', 'foreach', 'when', 'then', 'else', 'times', 'in',
                          'input', 'to', 'with', 'by', 'recipe', 'return', 'returns',
                          'cups', 'tbsp', 'tsp', 'grams', 'ml', 'oz', 'lbs',
                          'F', 'C', 'minutes', 'seconds', 'hours'])
    
    def compute_nullable(self):
        """Compute NULLABLE set for all non-terminals"""
        # Initialize all as False
        for nt in self.non_terminals:
            self.nullable[nt] = False
        
        # Iterate until no changes
        changed = True
        while changed:
            changed = False
            for lhs, productions in self.grammar.items():
                if self.nullable[lhs]:
                    continue
                    
                for rhs in productions:
                    # If production is epsilon, nullable
                    if len(rhs) == 1 and rhs[0] == self.epsilon:
                        self.nullable[lhs] = True
                        changed = True
                        break
                    
                    # If all symbols in rhs are nullable, lhs is nullable
                    all_nullable = True
                    for symbol in rhs:
                        if self.is_terminal(symbol):
                            all_nullable = False
                            break
                        if symbol in self.non_terminals and not self.nullable.get(symbol, False):
                            all_nullable = False
                            break
                    
                    if all_nullable:
                        self.nullable[lhs] = True
                        changed = True
                        break
    
    def compute_first(self):
        """Compute FIRST sets for all non-terminals"""
        # Initialize FIRST sets
        for nt in self.non_terminals:
            self.first[nt] = set()
        
        # Iterate until no changes
        changed = True
        while changed:
            changed = False
            
            for lhs, productions in self.grammar.items():
                for rhs in productions:
                    # If production is epsilon
                    if len(rhs) == 1 and rhs[0] == self.epsilon:
                        if self.epsilon not in self.first[lhs]:
                            self.first[lhs].add(self.epsilon)
                            changed = True
                        continue
                    
                    # Process each symbol in rhs
                    for i, symbol in enumerate(rhs):
                        if self.is_terminal(symbol):
                            # Add terminal to FIRST(lhs)
                            if symbol not in self.first[lhs]:
                                self.first[lhs].add(symbol)
                                changed = True
                            break
                        else:
                            # Add FIRST(symbol) - {ε} to FIRST(lhs)
                            before_size = len(self.first[lhs])
                            self.first[lhs] |= (self.first[symbol] - {self.epsilon})
                            if len(self.first[lhs]) > before_size:
                                changed = True
                            
                            # If symbol is not nullable, stop
                            if not self.nullable.get(symbol, False):
                                break
                            
                            # If all symbols are nullable, add epsilon
                            if i == len(rhs) - 1:
                                if self.epsilon not in self.first[lhs]:
                                    self.first[lhs].add(self.epsilon)
                                    changed = True
    
    def compute_follow(self, start_symbol):
        """Compute FOLLOW sets for all non-terminals"""
        # Initialize FOLLOW sets
        for nt in self.non_terminals:
            self.follow[nt] = set()
        
        # Add $ to FOLLOW(start_symbol)
        self.follow[start_symbol].add(self.end_marker)
        
        # Iterate until no changes
        changed = True
        while changed:
            changed = False
            
            for lhs, productions in self.grammar.items():
                for rhs in productions:
                    # Skip epsilon productions
                    if len(rhs) == 1 and rhs[0] == self.epsilon:
                        continue
                    
                    for i, symbol in enumerate(rhs):
                        if symbol not in self.non_terminals:
                            continue
                        
                        # Look at symbols after current symbol
                        rest = rhs[i+1:]
                        
                        if not rest:
                            # Symbol is at the end, add FOLLOW(lhs) to FOLLOW(symbol)
                            before_size = len(self.follow[symbol])
                            self.follow[symbol] |= self.follow[lhs]
                            if len(self.follow[symbol]) > before_size:
                                changed = True
                        else:
                            # Compute FIRST of rest
                            first_rest = self.compute_first_of_string(rest)
                            
                            # Add FIRST(rest) - {ε} to FOLLOW(symbol)
                            before_size = len(self.follow[symbol])
                            self.follow[symbol] |= (first_rest - {self.epsilon})
                            if len(self.follow[symbol]) > before_size:
                                changed = True
                            
                            # If rest is nullable, add FOLLOW(lhs) to FOLLOW(symbol)
                            if self.is_string_nullable(rest):
                                before_size = len(self.follow[symbol])
                                self.follow[symbol] |= self.follow[lhs]
                                if len(self.follow[symbol]) > before_size:
                                    changed = True
    
    def compute_first_of_string(self, symbols):
        """Compute FIRST set of a string of symbols"""
        result = set()
        
        for symbol in symbols:
            if self.is_terminal(symbol):
                result.add(symbol)
                return result
            else:
                result |= (self.first[symbol] - {self.epsilon})
                if not self.nullable.get(symbol, False):
                    return result
        
        # All symbols are nullable
        result.add(self.epsilon)
        return result
    
    def is_string_nullable(self, symbols):
        """Check if a string of symbols is nullable"""
        for symbol in symbols:
            if self.is_terminal(symbol):
                return False
            if not self.nullable.get(symbol, False):
                return False
        return True
    
    def build_parsing_table(self):
        """Build LL(1) predictive parsing table"""
        # Get all terminals
        self.terminals = set()
        for productions in self.grammar.values():
            for rhs in productions:
                for symbol in rhs:
                    if self.is_terminal(symbol) and symbol != self.epsilon:
                        self.terminals.add(symbol)
        
        # Add end marker
        self.terminals.add(self.end_marker)
        
        # Initialize parsing table
        for nt in self.non_terminals:
            self.parsing_table[nt] = {}
            for t in self.terminals:
                self.parsing_table[nt][t] = []
        
        # Fill parsing table
        for lhs, productions in self.grammar.items():
            for rhs in productions:
                # Check if this is an epsilon production
                is_epsilon_production = (len(rhs) == 1 and rhs[0] == self.epsilon)
                
                if is_epsilon_production:
                    # For epsilon productions, add to FOLLOW(lhs)
                    for terminal in self.follow[lhs]:
                        if terminal in self.terminals:
                            self.parsing_table[lhs][terminal].append(rhs)
                else:
                    # Compute FIRST of rhs
                    first_rhs = self.compute_first_of_string(rhs)
                    
                    # For each terminal in FIRST(rhs), add production to table
                    for terminal in first_rhs:
                        if terminal != self.epsilon and terminal in self.terminals:
                            self.parsing_table[lhs][terminal].append(rhs)
                    
                    # If rhs can derive epsilon, add production for FOLLOW(lhs)
                    if self.epsilon in first_rhs or self.is_string_nullable(rhs):
                        for terminal in self.follow[lhs]:
                            if terminal in self.terminals:
                                self.parsing_table[lhs][terminal].append(rhs)
    
    def validate_ll1(self):
        """Validate that the grammar is LL(1) and check for conflicts"""
        conflicts = []
        
        for nt in self.non_terminals:
            for terminal in self.terminals:
                productions = self.parsing_table[nt].get(terminal, [])
                if len(productions) > 1:
                    conflicts.append({
                        'non_terminal': nt,
                        'terminal': terminal,
                        'productions': productions
                    })
        
        return conflicts
    
    def export_to_excel(self, filename='ll1_parsing_table.xlsx'):
        """Export all results to Excel file"""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self.create_nullable_sheet(wb)
        self.create_first_sheet(wb)
        self.create_follow_sheet(wb)
        self.create_parsing_table_sheet(wb)
        self.create_grammar_sheet(wb)
        
        # Save workbook
        wb.save(filename)
        print(f"✓ Excel file saved: {filename}")
    
    def create_nullable_sheet(self, wb):
        """Create NULLABLE sheet"""
        ws = wb.create_sheet("NULLABLE")
        
        # Header
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "NULLABLE"
        
        # Style header
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 2
        for nt in sorted(self.non_terminals):
            ws[f'A{row}'] = nt
            ws[f'B{row}'] = "Yes" if self.nullable.get(nt, False) else "No"
            
            # Color code
            if self.nullable.get(nt, False):
                ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def create_first_sheet(self, wb):
        """Create FIRST sheet"""
        ws = wb.create_sheet("FIRST")
        
        # Header
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FIRST Set"
        
        # Style header
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 2
        for nt in sorted(self.non_terminals):
            ws[f'A{row}'] = nt
            first_set = sorted(self.first[nt])
            ws[f'B{row}'] = "{ " + ", ".join(first_set) + " }"
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
    
    def create_follow_sheet(self, wb):
        """Create FOLLOW sheet"""
        ws = wb.create_sheet("FOLLOW")
        
        # Header
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        
        # Style header
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="000000")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 2
        for nt in sorted(self.non_terminals):
            ws[f'A{row}'] = nt
            follow_set = sorted(self.follow[nt])
            ws[f'B{row}'] = "{ " + ", ".join(follow_set) + " }"
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
    
    def create_parsing_table_sheet(self, wb):
        """Create Predictive Parsing Table sheet"""
        ws = wb.create_sheet("Parsing Table")
        
        # Get sorted terminals and non-terminals
        terminals = sorted(self.terminals)
        non_terminals = sorted(self.non_terminals)
        
        # Header row
        ws['A1'] = "Non-Terminal"
        for col, terminal in enumerate(terminals, start=2):
            ws.cell(row=1, column=col, value=terminal)
        
        # Style header
        for col in range(1, len(terminals) + 2):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row, nt in enumerate(non_terminals, start=2):
            ws.cell(row=row, column=1, value=nt)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for col, terminal in enumerate(terminals, start=2):
                productions = self.parsing_table[nt].get(terminal, [])
                if productions:
                    # Format production
                    prod_strs = []
                    for prod in productions:
                        # Replace epsilon with ε symbol for display
                        prod_str = " ".join(prod).replace(self.epsilon, "ε")
                        prod_strs.append(f"{nt} → {prod_str}")
                    
                    cell_value = "\n".join(prod_strs)
                    cell = ws.cell(row=row, column=col, value=cell_value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                else:
                    cell = ws.cell(row=row, column=col, value="")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(terminals) + 2):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 25
        
        # Set row heights
        for row in range(2, len(non_terminals) + 2):
            ws.row_dimensions[row].height = 35
    
    def create_grammar_sheet(self, wb):
        """Create Grammar sheet"""
        ws = wb.create_sheet("Grammar", 0)  # Insert as first sheet
        
        # Header
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "Productions"
        
        # Style header
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = 2
        for lhs in sorted(self.grammar.keys()):
            ws[f'A{row}'] = lhs
            ws[f'A{row}'].font = Font(bold=True)
            
            productions = self.grammar[lhs]
            prod_strs = []
            for prod in productions:
                prod_str = " ".join(prod)
                prod_strs.append(prod_str)
            
            ws[f'B{row}'] = " | ".join(prod_strs)
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80


def define_recipescript_grammar():
    """Define RecipeScript grammar"""
    parser = LL1ParserGenerator()
    
    # Program Structure - Fixed to avoid conflicts
    parser.add_production('<program>', [
        ['<recipe_list>', '<statement_list>'],
        ['<statement_list>']
    ])
    
    parser.add_production('<recipe_list>', [
        ['<recipe_decl>', '<recipe_list_prime>']
    ])
    
    parser.add_production('<recipe_list_prime>', [
        ['<recipe_decl>', '<recipe_list_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<statement_list>', [
        ['<statement>', '<statement_list_prime>']
    ])
    
    parser.add_production('<statement_list_prime>', [
        ['<statement>', '<statement_list_prime>'],
        [parser.epsilon]
    ])
    
    # Recipe Functions - Fixed to avoid conflicts
    parser.add_production('<recipe_decl>', [
        ['recipe', 'IDENTIFIER', '(', '<param_list>', ')', '<recipe_return>', '{', '<statement_list>', '}']
    ])
    
    parser.add_production('<recipe_return>', [
        ['returns', '<type>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<param_list>', [
        ['<parameter>', '<param_list_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<param_list_prime>', [
        [',', '<parameter>', '<param_list_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<parameter>', [
        ['<type>', 'IDENTIFIER']
    ])
    
    parser.add_production('<recipe_call>', [
        ['IDENTIFIER', '(', '<arg_list>', ')']
    ])
    
    parser.add_production('<arg_list>', [
        ['<expression>', '<arg_list_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<arg_list_prime>', [
        [',', '<expression>', '<arg_list_prime>'],
        [parser.epsilon]
    ])
    
    # Statements
    parser.add_production('<statement>', [
        ['<input_stmt>', ';'],
        ['<declaration>', ';'],
        ['<operation>', ';'],
        ['<control_flow>'],
        ['<recipe_call>', ';'],
        ['<return_stmt>', ';']
    ])
    
    parser.add_production('<return_stmt>', [
        ['return', '<return_value>']
    ])
    
    parser.add_production('<return_value>', [
        ['<expression>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<input_stmt>', [
        ['input', 'IDENTIFIER']
    ])
    
    # Declarations
    parser.add_production('<declaration>', [
        ['<type>', 'IDENTIFIER', '=', '<value>']
    ])
    
    parser.add_production('<type>', [
        ['ingredient'],
        ['time'],
        ['temp'],
        ['quantity'],
        ['text']
    ])
    
    parser.add_production('<value>', [
        ['<expression>', '<value_tail>'],
        ['STRING']
    ])
    
    parser.add_production('<value_tail>', [
        ['<unit>'],
        [parser.epsilon]
    ])
    
    # Units
    parser.add_production('<unit>', [
        ['cups'], ['tbsp'], ['tsp'], ['grams'], ['ml'], ['oz'], ['lbs'],
        ['F'], ['C'],
        ['minutes'], ['seconds'], ['hours']
    ])
    
    # Operations
    parser.add_production('<operation>', [
        ['mix', '<ingredient_list>'],
        ['heat', 'IDENTIFIER', 'to', '<value>'],
        ['wait', '<value>'],
        ['serve', 'STRING'],
        ['display', 'IDENTIFIER'],
        ['add', 'IDENTIFIER', 'to', 'IDENTIFIER'],
        ['scale', 'IDENTIFIER', 'by', 'NUMBER']
    ])
    
    parser.add_production('<ingredient_list>', [
        ['IDENTIFIER', '<ingredient_list_prime>']
    ])
    
    parser.add_production('<ingredient_list_prime>', [
        ['with', 'IDENTIFIER', '<ingredient_list_prime>'],
        [parser.epsilon]
    ])
    
    # Control Flow
    parser.add_production('<control_flow>', [
        ['<repeat_stmt>'],
        ['<foreach_stmt>'],
        ['<when_stmt>']
    ])
    
    parser.add_production('<repeat_stmt>', [
        ['repeat', 'NUMBER', 'times', '{', '<statement_list>', '}']
    ])
    
    parser.add_production('<foreach_stmt>', [
        ['foreach', 'IDENTIFIER', 'in', 'IDENTIFIER', '{', '<statement_list>', '}']
    ])
    
    parser.add_production('<when_stmt>', [
        ['when', '<condition>', 'then', '{', '<statement_list>', '}', '<when_tail>']
    ])
    
    parser.add_production('<when_tail>', [
        ['else', '{', '<statement_list>', '}'],
        [parser.epsilon]
    ])
    
    # Conditions & Expressions
    parser.add_production('<condition>', [
        ['<expression>', '<comparison_op>', '<expression>']
    ])
    
    parser.add_production('<comparison_op>', [
        ['=='], ['!='], ['>'], ['<'], ['>='], ['<=']
    ])
    
    parser.add_production('<expression>', [
        ['<term>', '<expression_prime>']
    ])
    
    parser.add_production('<expression_prime>', [
        ['+', '<term>', '<expression_prime>'],
        ['-', '<term>', '<expression_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<term>', [
        ['<factor>', '<term_prime>']
    ])
    
    parser.add_production('<term_prime>', [
        ['*', '<factor>', '<term_prime>'],
        ['/', '<factor>', '<term_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<factor>', [
        ['NUMBER'],
        ['IDENTIFIER', '<factor_tail>'],
        ['(', '<expression>', ')']
    ])
    
    parser.add_production('<factor_tail>', [
        ['(', '<arg_list>', ')'],
        [parser.epsilon]
    ])
    
    return parser


if __name__ == "__main__":
    print("=" * 60)
    print("LL(1) PREDICTIVE PARSER GENERATOR")
    print("RecipeScript Compiler")
    print("=" * 60)
    
    # Define grammar
    print("\n[1/5] Defining grammar...")
    parser = define_recipescript_grammar()
    print(f"✓ Grammar defined: {len(parser.grammar)} non-terminals")
    
    # Compute NULLABLE
    print("\n[2/5] Computing NULLABLE sets...")
    parser.compute_nullable()
    nullable_count = sum(1 for v in parser.nullable.values() if v)
    print(f"✓ NULLABLE computed: {nullable_count} nullable non-terminals")
    
    # Compute FIRST
    print("\n[3/5] Computing FIRST sets...")
    parser.compute_first()
    print(f"✓ FIRST computed for {len(parser.first)} non-terminals")
    
    # Compute FOLLOW
    print("\n[4/5] Computing FOLLOW sets...")
    parser.compute_follow('<program>')
    print(f"✓ FOLLOW computed for {len(parser.follow)} non-terminals")
    
    # Build parsing table
    print("\n[5/5] Building predictive parsing table...")
    parser.build_parsing_table()
    print(f"✓ Parsing table built: {len(parser.terminals)} terminals")
    
    # Validate LL(1)
    print("\n[6/6] Validating LL(1) grammar...")
    conflicts = parser.validate_ll1()
    if conflicts:
        print(f"⚠ WARNING: Found {len(conflicts)} conflicts!")
        for conflict in conflicts[:5]:  # Show first 5
            print(f"  - {conflict['non_terminal']} on '{conflict['terminal']}': {len(conflict['productions'])} productions")
    else:
        print("✓ No conflicts found - Grammar is LL(1)!")
    
    # Export to Excel
    print("\n" + "=" * 60)
    print("EXPORTING TO EXCEL")
    print("=" * 60)
    parser.export_to_excel('ll1_parsing_table.xlsx')
    
    print("\n" + "=" * 60)
    print("✓ COMPLETE!")
    print("=" * 60)
    print("\nGenerated file: ll1_parsing_table.xlsx")
    print("\nSheets included:")
    print("  1. Grammar - Complete grammar rules")
    print("  2. NULLABLE - Nullable non-terminals")
    print("  3. FIRST - FIRST sets for all non-terminals")
    print("  4. FOLLOW - FOLLOW sets for all non-terminals")
    print("  5. Parsing Table - LL(1) predictive parsing table")
    
    if conflicts:
        print(f"\n⚠ Note: {len(conflicts)} conflicts detected")
        print("  Check the parsing table for cells with multiple productions")
    else:
        print("\n✓ Grammar is valid LL(1) - No conflicts!")
    
    print("\n" + "=" * 60)
