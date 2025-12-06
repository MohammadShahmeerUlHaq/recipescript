"""
LL(1) Predictive Parser Generator - MINI GRAMMAR for Chart Paper
Simplified RecipeScript grammar with 5-7 key productions
"""

import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class LL1ParserGenerator:
    def __init__(self):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.nullable = {}
        self.first = defaultdict(set)
        self.follow = defaultdict(set)
        self.parsing_table = {}
        self.epsilon = 'ε'
        self.end_marker = '$'
        
    def add_production(self, lhs, rhs_list):
        """Add a production rule: lhs -> rhs1 | rhs2 | ..."""
        if lhs not in self.grammar:
            self.grammar[lhs] = []
        self.non_terminals.add(lhs)
        
        for rhs in rhs_list:
            self.grammar[lhs].append(rhs)
            for symbol in rhs:
                if symbol != self.epsilon and not self.is_terminal(symbol):
                    self.non_terminals.add(symbol)
    
    def is_terminal(self, symbol):
        """Check if symbol is terminal"""
        if symbol == self.epsilon or symbol == self.end_marker:
            return False
        return (symbol.isupper() or 
                symbol in ['+', '-', '*', '/', '=', '==', '>', '<', ';', ',', '(', ')'] or
                symbol in ['ingredient', 'mix', 'when', 'then', 'cups'])
    
    def compute_nullable(self):
        """Compute NULLABLE set for all non-terminals"""
        for nt in self.non_terminals:
            self.nullable[nt] = False
        
        changed = True
        while changed:
            changed = False
            for lhs, productions in self.grammar.items():
                if self.nullable[lhs]:
                    continue
                    
                for rhs in productions:
                    if len(rhs) == 1 and rhs[0] == self.epsilon:
                        self.nullable[lhs] = True
                        changed = True
                        break
                    
                    all_nullable = True
                    for symbol in rhs:
                        if self.is_terminal(symbol):
                            all_nullable = False
                            break
                        if symbol in self.non_terminals and not self.nullable.get(symbol, False):
                            all_nullable = False
                            break
                    
                    if all_nullable:
                        self.nullable[lhs] = True
                        changed = True
                        break
    
    def compute_first(self):
        """Compute FIRST sets for all non-terminals"""
        for nt in self.non_terminals:
            self.first[nt] = set()
        
        changed = True
        while changed:
            changed = False
            
            for lhs, productions in self.grammar.items():
                for rhs in productions:
                    if len(rhs) == 1 and rhs[0] == self.epsilon:
                        if self.epsilon not in self.first[lhs]:
                            self.first[lhs].add(self.epsilon)
                            changed = True
                        continue
                    
                    for i, symbol in enumerate(rhs):
                        if self.is_terminal(symbol):
                            if symbol not in self.first[lhs]:
                                self.first[lhs].add(symbol)
                                changed = True
                            break
                        else:
                            before_size = len(self.first[lhs])
                            self.first[lhs] |= (self.first[symbol] - {self.epsilon})
                            if len(self.first[lhs]) > before_size:
                                changed = True
                            
                            if not self.nullable.get(symbol, False):
                                break
                            
                            if i == len(rhs) - 1:
                                if self.epsilon not in self.first[lhs]:
                                    self.first[lhs].add(self.epsilon)
                                    changed = True
    
    def compute_follow(self, start_symbol):
        """Compute FOLLOW sets for all non-terminals"""
        for nt in self.non_terminals:
            self.follow[nt] = set()
        
        self.follow[start_symbol].add(self.end_marker)
        
        changed = True
        while changed:
            changed = False
            
            for lhs, productions in self.grammar.items():
                for rhs in productions:
                    if len(rhs) == 1 and rhs[0] == self.epsilon:
                        continue
                    
                    for i, symbol in enumerate(rhs):
                        if symbol not in self.non_terminals:
                            continue
                        
                        rest = rhs[i+1:]
                        
                        if not rest:
                            before_size = len(self.follow[symbol])
                            self.follow[symbol] |= self.follow[lhs]
                            if len(self.follow[symbol]) > before_size:
                                changed = True
                        else:
                            first_rest = self.compute_first_of_string(rest)
                            
                            before_size = len(self.follow[symbol])
                            self.follow[symbol] |= (first_rest - {self.epsilon})
                            if len(self.follow[symbol]) > before_size:
                                changed = True
                            
                            if self.is_string_nullable(rest):
                                before_size = len(self.follow[symbol])
                                self.follow[symbol] |= self.follow[lhs]
                                if len(self.follow[symbol]) > before_size:
                                    changed = True
    
    def compute_first_of_string(self, symbols):
        """Compute FIRST set of a string of symbols"""
        result = set()
        
        for symbol in symbols:
            if self.is_terminal(symbol):
                result.add(symbol)
                return result
            else:
                result |= (self.first[symbol] - {self.epsilon})
                if not self.nullable.get(symbol, False):
                    return result
        
        result.add(self.epsilon)
        return result
    
    def is_string_nullable(self, symbols):
        """Check if a string of symbols is nullable"""
        for symbol in symbols:
            if self.is_terminal(symbol):
                return False
            if not self.nullable.get(symbol, False):
                return False
        return True
    
    def build_parsing_table(self):
        """Build LL(1) predictive parsing table"""
        self.terminals = set()
        for productions in self.grammar.values():
            for rhs in productions:
                for symbol in rhs:
                    if self.is_terminal(symbol) and symbol != self.epsilon:
                        self.terminals.add(symbol)
        
        self.terminals.add(self.end_marker)
        
        for nt in self.non_terminals:
            self.parsing_table[nt] = {}
            for t in self.terminals:
                self.parsing_table[nt][t] = []
        
        for lhs, productions in self.grammar.items():
            for rhs in productions:
                is_epsilon_production = (len(rhs) == 1 and rhs[0] == self.epsilon)
                
                if is_epsilon_production:
                    for terminal in self.follow[lhs]:
                        if terminal in self.terminals:
                            self.parsing_table[lhs][terminal].append(rhs)
                else:
                    first_rhs = self.compute_first_of_string(rhs)
                    
                    for terminal in first_rhs:
                        if terminal != self.epsilon and terminal in self.terminals:
                            self.parsing_table[lhs][terminal].append(rhs)
                    
                    if self.epsilon in first_rhs or self.is_string_nullable(rhs):
                        for terminal in self.follow[lhs]:
                            if terminal in self.terminals:
                                self.parsing_table[lhs][terminal].append(rhs)
    
    def validate_ll1(self):
        """Validate that the grammar is LL(1)"""
        conflicts = []
        
        for nt in self.non_terminals:
            for terminal in self.terminals:
                productions = self.parsing_table[nt].get(terminal, [])
                if len(productions) > 1:
                    conflicts.append({
                        'non_terminal': nt,
                        'terminal': terminal,
                        'productions': productions
                    })
        
        return conflicts
    
    def export_to_excel(self, filename='ll1_mini_parsing_table.xlsx'):
        """Export all results to Excel file"""
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        self.create_grammar_sheet(wb)
        self.create_nullable_sheet(wb)
        self.create_first_sheet(wb)
        self.create_follow_sheet(wb)
        self.create_parsing_table_sheet(wb)
        
        wb.save(filename)
        print(f"✓ Excel file saved: {filename}")
    
    def create_grammar_sheet(self, wb):
        """Create Grammar sheet"""
        ws = wb.create_sheet("Grammar", 0)
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "Productions"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        for lhs in sorted(self.grammar.keys()):
            ws[f'A{row}'] = lhs
            ws[f'A{row}'].font = Font(bold=True)
            
            productions = self.grammar[lhs]
            prod_strs = []
            for prod in productions:
                prod_str = " ".join(prod)
                prod_strs.append(prod_str)
            
            ws[f'B{row}'] = " | ".join(prod_strs)
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
    
    def create_nullable_sheet(self, wb):
        """Create NULLABLE sheet"""
        ws = wb.create_sheet("NULLABLE")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "NULLABLE"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        for nt in sorted(self.non_terminals):
            ws[f'A{row}'] = nt
            ws[f'B{row}'] = "Yes" if self.nullable.get(nt, False) else "No"
            
            if self.nullable.get(nt, False):
                ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def create_first_sheet(self, wb):
        """Create FIRST sheet"""
        ws = wb.create_sheet("FIRST")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FIRST Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        for nt in sorted(self.non_terminals):
            ws[f'A{row}'] = nt
            first_set = sorted(self.first[nt])
            ws[f'B{row}'] = "{ " + ", ".join(first_set) + " }"
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def create_follow_sheet(self, wb):
        """Create FOLLOW sheet"""
        ws = wb.create_sheet("FOLLOW")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="000000")
            ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        row = 2
        for nt in sorted(self.non_terminals):
            ws[f'A{row}'] = nt
            follow_set = sorted(self.follow[nt])
            ws[f'B{row}'] = "{ " + ", ".join(follow_set) + " }"
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def create_parsing_table_sheet(self, wb):
        """Create Predictive Parsing Table sheet"""
        ws = wb.create_sheet("Parsing Table")
        
        terminals = sorted(self.terminals)
        non_terminals = sorted(self.non_terminals)
        
        ws['A1'] = "Non-Terminal"
        for col, terminal in enumerate(terminals, start=2):
            ws.cell(row=1, column=col, value=terminal)
        
        for col in range(1, len(terminals) + 2):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row, nt in enumerate(non_terminals, start=2):
            ws.cell(row=row, column=1, value=nt)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for col, terminal in enumerate(terminals, start=2):
                productions = self.parsing_table[nt].get(terminal, [])
                if productions:
                    prod_strs = []
                    for prod in productions:
                        prod_str = " ".join(prod).replace(self.epsilon, "ε")
                        prod_strs.append(f"{nt} → {prod_str}")
                    
                    cell_value = "\n".join(prod_strs)
                    cell = ws.cell(row=row, column=col, value=cell_value)
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                else:
                    cell = ws.cell(row=row, column=col, value="")
        
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(terminals) + 2):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22
        
        for row in range(2, len(non_terminals) + 2):
            ws.row_dimensions[row].height = 35


def define_mini_grammar():
    """
    Define MINI RecipeScript grammar for chart paper
    
    Covers:
    1. Variable declaration (ingredient flour = 2 cups;)
    2. Arithmetic expressions (2 + 3 * 4)
    3. Conditional statement (when flour > 2 then { mix flour; })
    
    Total: 7 non-terminals, ~15 productions
    """
    parser = LL1ParserGenerator()
    
    # 1. Program and Statement
    parser.add_production('<program>', [
        ['<statement>']
    ])
    
    parser.add_production('<statement>', [
        ['<declaration>', ';'],
        ['<operation>', ';'],
        ['<when_stmt>']
    ])
    
    # 2. Declaration (ingredient flour = 2 cups;)
    parser.add_production('<declaration>', [
        ['ingredient', 'IDENTIFIER', '=', '<expression>', '<unit_opt>']
    ])
    
    parser.add_production('<unit_opt>', [
        ['cups'],
        [parser.epsilon]
    ])
    
    # 3. Operation (mix flour;)
    parser.add_production('<operation>', [
        ['mix', 'IDENTIFIER']
    ])
    
    # 4. Conditional (when flour > 2 then { mix flour; })
    parser.add_production('<when_stmt>', [
        ['when', '<condition>', 'then', '<statement>']
    ])
    
    parser.add_production('<condition>', [
        ['<expression>', '>', '<expression>']
    ])
    
    # 5. Expressions (2 + 3 * 4)
    parser.add_production('<expression>', [
        ['<term>', '<expr_prime>']
    ])
    
    parser.add_production('<expr_prime>', [
        ['+', '<term>', '<expr_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<term>', [
        ['<factor>', '<term_prime>']
    ])
    
    parser.add_production('<term_prime>', [
        ['*', '<factor>', '<term_prime>'],
        [parser.epsilon]
    ])
    
    parser.add_production('<factor>', [
        ['NUMBER'],
        ['IDENTIFIER']
    ])
    
    return parser


if __name__ == "__main__":
    print("=" * 70)
    print("LL(1) PREDICTIVE PARSER - MINI GRAMMAR FOR CHART PAPER")
    print("RecipeScript Simplified Grammar")
    print("=" * 70)
    
    print("\n[1/5] Defining mini grammar...")
    parser = define_mini_grammar()
    print(f"✓ Grammar defined: {len(parser.grammar)} non-terminals")
    print(f"  Productions cover:")
    print(f"    - Variable declarations (ingredient flour = 2 cups;)")
    print(f"    - Arithmetic expressions (2 + 3 * 4)")
    print(f"    - Conditional statements (when flour > 2 then ...)")
    
    print("\n[2/5] Computing NULLABLE sets...")
    parser.compute_nullable()
    nullable_count = sum(1 for v in parser.nullable.values() if v)
    print(f"✓ NULLABLE computed: {nullable_count} nullable non-terminals")
    
    print("\n[3/5] Computing FIRST sets...")
    parser.compute_first()
    print(f"✓ FIRST computed for {len(parser.first)} non-terminals")
    
    print("\n[4/5] Computing FOLLOW sets...")
    parser.compute_follow('<program>')
    print(f"✓ FOLLOW computed for {len(parser.follow)} non-terminals")
    
    print("\n[5/5] Building predictive parsing table...")
    parser.build_parsing_table()
    print(f"✓ Parsing table built: {len(parser.terminals)} terminals")
    
    print("\n[6/6] Validating LL(1) grammar...")
    conflicts = parser.validate_ll1()
    if conflicts:
        print(f"⚠ WARNING: Found {len(conflicts)} conflicts!")
    else:
        print("✓ No conflicts found - Grammar is LL(1)!")
    
    print("\n" + "=" * 70)
    print("EXPORTING TO EXCEL")
    print("=" * 70)
    parser.export_to_excel('ll1_mini_parsing_table.xlsx')
    
    print("\n" + "=" * 70)
    print("✓ COMPLETE!")
    print("=" * 70)
    print("\nGenerated file: ll1_mini_parsing_table.xlsx")
    print("\nSheets included:")
    print("  1. Grammar - 12 non-terminals, compact for chart paper")
    print("  2. NULLABLE - Easy to draw")
    print("  3. FIRST - Manageable size")
    print("  4. FOLLOW - Manageable size")
    print("  5. Parsing Table - Fits on one chart paper!")
    
    print("\n" + "=" * 70)
    print("GRAMMAR SUMMARY")
    print("=" * 70)
    print(f"Non-terminals: {len(parser.non_terminals)}")
    print(f"Terminals: {len(parser.terminals)}")
    print(f"Nullable: {nullable_count}")
    print(f"Parsing table size: {len(parser.non_terminals)} × {len(parser.terminals)}")
    print("\nThis grammar is PERFECT for chart paper! ✓")
    print("=" * 70)
