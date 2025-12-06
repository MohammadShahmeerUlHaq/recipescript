"""
LR(0) and SLR(1) Parser Generator for RecipeScript Mini Grammar
Generates complete Excel file with states, transitions, and ACTION/GOTO tables
"""

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class LR0Item:
    """LR(0) Item: [A → α·β] (no lookahead)"""
    def __init__(self, lhs, rhs, dot):
        self.lhs = lhs
        self.rhs = tuple(rhs)
        self.dot = dot
    
    def __eq__(self, other):
        return (self.lhs == other.lhs and self.rhs == other.rhs and self.dot == other.dot)
    
    def __hash__(self):
        return hash((self.lhs, self.rhs, self.dot))
    
    def __repr__(self):
        rhs_list = list(self.rhs) if self.rhs else ['ε']
        rhs_with_dot = rhs_list[:self.dot] + ['•'] + rhs_list[self.dot:]
        return f"[{self.lhs} → {' '.join(rhs_with_dot)}]"
    
    def next_symbol(self):
        if self.dot < len(self.rhs):
            return self.rhs[self.dot]
        return None
    
    def is_complete(self):
        return self.dot >= len(self.rhs)
    
    def advance(self):
        return LR0Item(self.lhs, self.rhs, self.dot + 1)


class LR0_SLR_Parser:
    def __init__(self):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.start = None
        self.aug_start = None
        
        self.states = []
        self.state_map = {}
        self.transitions = {}
        
        self.first = defaultdict(set)
        self.follow = defaultdict(set)
        
        self.lr0_action = {}
        self.lr0_goto = {}
        self.lr0_conflicts = []
        
        self.slr_action = {}
        self.slr_goto = {}
        self.slr_conflicts = []
    
    def add_production(self, lhs, rhs_list):
        if lhs not in self.grammar:
            self.grammar[lhs] = []
        self.non_terminals.add(lhs)
        for rhs in rhs_list:
            self.grammar[lhs].append(tuple(rhs))
    
    def is_terminal(self, sym):
        return sym not in self.non_terminals and sym != '$'
    
    def augment_grammar(self, start_symbol):
        self.start = start_symbol
        self.aug_start = start_symbol + "'"
        self.grammar[self.aug_start] = [(start_symbol,)]
        self.non_terminals.add(self.aug_start)
    
    def compute_first(self):
        """Compute FIRST sets (needed for completeness)"""
        for nt in self.non_terminals:
            self.first[nt] = set()
        
        changed = True
        while changed:
            changed = False
            for lhs, prods in self.grammar.items():
                for rhs in prods:
                    if not rhs:
                        if 'ε' not in self.first[lhs]:
                            self.first[lhs].add('ε')
                            changed = True
                    else:
                        for sym in rhs:
                            if self.is_terminal(sym):
                                if sym not in self.first[lhs]:
                                    self.first[lhs].add(sym)
                                    changed = True
                                break
                            else:
                                old_size = len(self.first[lhs])
                                self.first[lhs] |= (self.first[sym] - {'ε'})
                                if len(self.first[lhs]) > old_size:
                                    changed = True
                                if 'ε' not in self.first[sym]:
                                    break
    
    def compute_follow(self):
        """Compute FOLLOW sets (needed for SLR)"""
        for nt in self.non_terminals:
            self.follow[nt] = set()
        
        self.follow[self.start].add('$')
        
        changed = True
        while changed:
            changed = False
            for lhs, prods in self.grammar.items():
                for rhs in prods:
                    for i, sym in enumerate(rhs):
                        if sym not in self.non_terminals:
                            continue
                        
                        rest = rhs[i+1:]
                        if not rest:
                            old_size = len(self.follow[sym])
                            self.follow[sym] |= self.follow[lhs]
                            if len(self.follow[sym]) > old_size:
                                changed = True
                        else:
                            for next_sym in rest:
                                if self.is_terminal(next_sym):
                                    if next_sym not in self.follow[sym]:
                                        self.follow[sym].add(next_sym)
                                        changed = True
                                    break
                                else:
                                    old_size = len(self.follow[sym])
                                    self.follow[sym] |= (self.first[next_sym] - {'ε'})
                                    if len(self.follow[sym]) > old_size:
                                        changed = True
                                    if 'ε' not in self.first[next_sym]:
                                        break
                            else:
                                old_size = len(self.follow[sym])
                                self.follow[sym] |= self.follow[lhs]
                                if len(self.follow[sym]) > old_size:
                                    changed = True
    
    def closure(self, items):
        """Compute closure of LR(0) items"""
        closure_set = set(items)
        changed = True
        
        while changed:
            changed = False
            new_items = set()
            for item in closure_set:
                next_sym = item.next_symbol()
                if next_sym and next_sym in self.non_terminals:
                    for prod in self.grammar.get(next_sym, []):
                        new_item = LR0Item(next_sym, prod, 0)
                        if new_item not in closure_set:
                            new_items.add(new_item)
                            changed = True
            closure_set |= new_items
        
        return frozenset(closure_set)
    
    def goto(self, items, symbol):
        """Compute GOTO(items, symbol)"""
        moved = set()
        for item in items:
            if item.next_symbol() == symbol:
                moved.add(item.advance())
        
        if moved:
            return self.closure(moved)
        return frozenset()
    
    def build_automaton(self):
        """Build LR(0) automaton"""
        # Collect terminals
        for prods in self.grammar.values():
            for prod in prods:
                for sym in prod:
                    if self.is_terminal(sym):
                        self.terminals.add(sym)
        self.terminals.add('$')
        
        # Initial state
        initial_item = LR0Item(self.aug_start, self.grammar[self.aug_start][0], 0)
        initial_state = self.closure({initial_item})
        
        self.states = [initial_state]
        self.state_map[initial_state] = 0
        queue = [initial_state]
        
        while queue:
            current = queue.pop(0)
            current_idx = self.state_map[current]
            
            symbols = set()
            for item in current:
                sym = item.next_symbol()
                if sym:
                    symbols.add(sym)
            
            for sym in symbols:
                next_state = self.goto(current, sym)
                if next_state and next_state not in self.state_map:
                    self.state_map[next_state] = len(self.states)
                    self.states.append(next_state)
                    queue.append(next_state)
                
                if next_state:
                    self.transitions[(current_idx, sym)] = self.state_map[next_state]
        
        print(f"✓ Built {len(self.states)} states")
        print(f"✓ Built {len(self.transitions)} transitions")
    
    def get_production_number(self, lhs, rhs):
        """Get production number"""
        prod_num = 0
        for nt in sorted(self.grammar.keys()):
            for prod in self.grammar[nt]:
                if nt == lhs and prod == rhs:
                    return prod_num
                prod_num += 1
        return -1
    
    def build_lr0_tables(self):
        """Build LR(0) parsing tables"""
        print("\n=== Building LR(0) Tables ===")
        
        for idx, state in enumerate(self.states):
            self.lr0_action[idx] = {}
            self.lr0_goto[idx] = {}
            
            for t in self.terminals:
                self.lr0_action[idx][t] = []
            
            for nt in self.non_terminals:
                if nt != self.aug_start:
                    self.lr0_goto[idx][nt] = None
        
        # Fill tables
        for idx, state in enumerate(self.states):
            for item in state:
                if item.is_complete():
                    if item.lhs == self.aug_start:
                        self.lr0_action[idx]['$'].append(('accept',))
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        # LR(0): Add reduce for ALL terminals
                        for t in self.terminals:
                            self.lr0_action[idx][t].append(('reduce', prod_num))
                else:
                    sym = item.next_symbol()
                    if sym and self.is_terminal(sym):
                        if (idx, sym) in self.transitions:
                            next_idx = self.transitions[(idx, sym)]
                            shift_action = ('shift', next_idx)
                            if shift_action not in self.lr0_action[idx][sym]:
                                self.lr0_action[idx][sym].append(shift_action)
            
            # GOTO
            for nt in self.non_terminals:
                if nt != self.aug_start and (idx, nt) in self.transitions:
                    self.lr0_goto[idx][nt] = self.transitions[(idx, nt)]
        
        # Check conflicts
        for idx in range(len(self.states)):
            for t in self.terminals:
                actions = self.lr0_action[idx][t]
                if len(actions) > 1:
                    self.lr0_conflicts.append((idx, t, actions))
        
        if self.lr0_conflicts:
            print(f"⚠ LR(0): Found {len(self.lr0_conflicts)} conflicts")
        else:
            print("✓ LR(0): No conflicts!")
    
    def build_slr_tables(self):
        """Build SLR(1) parsing tables"""
        print("\n=== Building SLR(1) Tables ===")
        
        for idx, state in enumerate(self.states):
            self.slr_action[idx] = {}
            self.slr_goto[idx] = {}
            
            for t in self.terminals:
                self.slr_action[idx][t] = []
            
            for nt in self.non_terminals:
                if nt != self.aug_start:
                    self.slr_goto[idx][nt] = None
        
        # Fill tables
        for idx, state in enumerate(self.states):
            for item in state:
                if item.is_complete():
                    if item.lhs == self.aug_start:
                        self.slr_action[idx]['$'].append(('accept',))
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        # SLR: Add reduce only for FOLLOW(lhs)
                        for t in self.follow[item.lhs]:
                            if t in self.terminals:
                                self.slr_action[idx][t].append(('reduce', prod_num))
                else:
                    sym = item.next_symbol()
                    if sym and self.is_terminal(sym):
                        if (idx, sym) in self.transitions:
                            next_idx = self.transitions[(idx, sym)]
                            shift_action = ('shift', next_idx)
                            if shift_action not in self.slr_action[idx][sym]:
                                self.slr_action[idx][sym].append(shift_action)
            
            # GOTO (same as LR(0))
            for nt in self.non_terminals:
                if nt != self.aug_start and (idx, nt) in self.transitions:
                    self.slr_goto[idx][nt] = self.transitions[(idx, nt)]
        
        # Check conflicts
        for idx in range(len(self.states)):
            for t in self.terminals:
                actions = self.slr_action[idx][t]
                if len(actions) > 1:
                    self.slr_conflicts.append((idx, t, actions))
        
        if self.slr_conflicts:
            print(f"⚠ SLR(1): Found {len(self.slr_conflicts)} conflicts")
        else:
            print("✓ SLR(1): No conflicts!")

    
    def export_to_excel(self, filename='lr0_slr_analysis.xlsx'):
        """Export complete analysis to Excel"""
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        self.create_summary_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_follow_sheet(wb)
        self.create_states_sheet(wb)
        self.create_transitions_sheet(wb)
        self.create_lr0_action_sheet(wb)
        self.create_lr0_goto_sheet(wb)
        self.create_slr_action_sheet(wb)
        self.create_slr_goto_sheet(wb)
        
        wb.save(filename)
        print(f"\n✓ Excel file saved: {filename}")
    
    def create_summary_sheet(self, wb):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        ws['A1'] = "LR(0) and SLR(1) Parser Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Grammar Statistics:"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A4'] = "Non-terminals:"
        ws['B4'] = len(self.non_terminals)
        ws['A5'] = "Terminals:"
        ws['B5'] = len(self.terminals)
        ws['A6'] = "States:"
        ws['B6'] = len(self.states)
        ws['A7'] = "Transitions:"
        ws['B7'] = len(self.transitions)
        
        ws['A9'] = "LR(0) Analysis:"
        ws['A9'].font = Font(bold=True, size=12, color="FF0000" if self.lr0_conflicts else "008000")
        
        ws['A10'] = "Is LR(0)?"
        ws['B10'] = "NO ✗" if self.lr0_conflicts else "YES ✓"
        ws['B10'].font = Font(bold=True, color="FF0000" if self.lr0_conflicts else "008000")
        
        ws['A11'] = "Conflicts:"
        ws['B11'] = len(self.lr0_conflicts)
        
        ws['A13'] = "SLR(1) Analysis:"
        ws['A13'].font = Font(bold=True, size=12, color="FF0000" if self.slr_conflicts else "008000")
        
        ws['A14'] = "Is SLR(1)?"
        ws['B14'] = "NO ✗" if self.slr_conflicts else "YES ✓"
        ws['B14'].font = Font(bold=True, color="FF0000" if self.slr_conflicts else "008000")
        
        ws['A15'] = "Conflicts:"
        ws['B15'] = len(self.slr_conflicts)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Prod #"
        ws['B1'] = "LHS"
        ws['C1'] = "RHS"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="4472C4", fill_type="solid")
        
        row = 2
        prod_num = 0
        for lhs in sorted(self.grammar.keys()):
            for rhs in self.grammar[lhs]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = lhs
                ws[f'C{row}'] = ' '.join(rhs) if rhs else 'ε'
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
    
    def create_follow_sheet(self, wb):
        """Create FOLLOW sets sheet"""
        ws = wb.create_sheet("FOLLOW Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="FFC000", fill_type="solid")
        
        row = 2
        for nt in sorted(self.non_terminals):
            if nt == self.aug_start:
                continue
            ws[f'A{row}'] = nt
            ws[f'B{row}'] = ', '.join(sorted(self.follow[nt]))
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def create_states_sheet(self, wb):
        """Create states sheet"""
        ws = wb.create_sheet("States")
        
        ws['A1'] = "State"
        ws['B1'] = "LR(0) Items"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="70AD47", fill_type="solid")
        
        for idx, state in enumerate(self.states):
            row = idx + 2
            ws[f'A{row}'] = f"I{idx}"
            ws[f'A{row}'].font = Font(bold=True)
            
            items_str = '\n'.join(str(item) for item in sorted(state, key=str))
            ws[f'B{row}'] = items_str
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = max(20, len(state) * 15)
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
    
    def create_transitions_sheet(self, wb):
        """Create transitions sheet"""
        ws = wb.create_sheet("Transitions")
        
        ws['A1'] = "From"
        ws['B1'] = "Symbol"
        ws['C1'] = "To"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="ED7D31", fill_type="solid")
        
        row = 2
        for (from_st, sym), to_st in sorted(self.transitions.items()):
            ws[f'A{row}'] = f"I{from_st}"
            ws[f'B{row}'] = sym
            ws[f'C{row}'] = f"I{to_st}"
            row += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
    
    def create_lr0_action_sheet(self, wb):
        """Create LR(0) ACTION table"""
        ws = wb.create_sheet("LR0 ACTION")
        
        terminals = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col, t in enumerate(terminals, 2):
            ws.cell(1, col, t)
        
        for col in range(1, len(terminals) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="4472C4", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, t in enumerate(terminals, 2):
                actions = self.lr0_action[idx][t]
                if actions:
                    strs = []
                    for act in actions:
                        if act[0] == 'shift':
                            strs.append(f"s{act[1]}")
                        elif act[0] == 'reduce':
                            strs.append(f"r{act[1]}")
                        elif act[0] == 'accept':
                            strs.append("acc")
                    
                    cell = ws.cell(row, col, '\n'.join(strs))
                    cell.alignment = Alignment(wrap_text=True, horizontal='center')
                    
                    if len(actions) > 1:
                        cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(terminals) + 2):
            ws.column_dimensions[chr(64+col)].width = 12
    
    def create_lr0_goto_sheet(self, wb):
        """Create LR(0) GOTO table"""
        ws = wb.create_sheet("LR0 GOTO")
        
        nts = sorted([nt for nt in self.non_terminals if nt != self.aug_start])
        
        ws['A1'] = "State"
        for col, nt in enumerate(nts, 2):
            ws.cell(1, col, nt)
        
        for col in range(1, len(nts) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="4472C4", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, nt in enumerate(nts, 2):
                val = self.lr0_goto[idx][nt]
                if val is not None:
                    cell = ws.cell(row, col, val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="D9EAD3", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(nts) + 2):
            ws.column_dimensions[chr(64+col)].width = 15
    
    def create_slr_action_sheet(self, wb):
        """Create SLR(1) ACTION table"""
        ws = wb.create_sheet("SLR ACTION")
        
        terminals = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col, t in enumerate(terminals, 2):
            ws.cell(1, col, t)
        
        for col in range(1, len(terminals) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="70AD47", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, t in enumerate(terminals, 2):
                actions = self.slr_action[idx][t]
                if actions:
                    strs = []
                    for act in actions:
                        if act[0] == 'shift':
                            strs.append(f"s{act[1]}")
                        elif act[0] == 'reduce':
                            strs.append(f"r{act[1]}")
                        elif act[0] == 'accept':
                            strs.append("acc")
                    
                    cell = ws.cell(row, col, '\n'.join(strs))
                    cell.alignment = Alignment(wrap_text=True, horizontal='center')
                    
                    if len(actions) > 1:
                        cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(terminals) + 2):
            ws.column_dimensions[chr(64+col)].width = 12
    
    def create_slr_goto_sheet(self, wb):
        """Create SLR(1) GOTO table"""
        ws = wb.create_sheet("SLR GOTO")
        
        nts = sorted([nt for nt in self.non_terminals if nt != self.aug_start])
        
        ws['A1'] = "State"
        for col, nt in enumerate(nts, 2):
            ws.cell(1, col, nt)
        
        for col in range(1, len(nts) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="70AD47", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, nt in enumerate(nts, 2):
                val = self.slr_goto[idx][nt]
                if val is not None:
                    cell = ws.cell(row, col, val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="D9EAD3", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(nts) + 2):
            ws.column_dimensions[chr(64+col)].width = 15


# Define mini grammar
parser = LR0_SLR_Parser()

# Productions
parser.add_production('S', [['D'], ['O']])
parser.add_production('D', [['ingredient', 'id', '=', 'E']])
parser.add_production('O', [['mix', 'id']])
parser.add_production('E', [['E', '+', 'T'], ['T']])
parser.add_production('T', [['num'], ['id']])

print("=" * 70)
print("LR(0) AND SLR(1) PARSER ANALYSIS")
print("RecipeScript Mini Grammar")
print("=" * 70)

print("\nGrammar:")
for lhs in sorted(parser.grammar.keys()):
    for rhs in parser.grammar[lhs]:
        print(f"  {lhs} → {' '.join(rhs)}")

print("\n[1/5] Augmenting grammar...")
parser.augment_grammar('S')
print(f"✓ Augmented start: {parser.aug_start}")

print("\n[2/5] Computing FIRST and FOLLOW sets...")
parser.compute_first()
parser.compute_follow()
print("✓ FIRST and FOLLOW computed")

print("\n[3/5] Building LR(0) automaton...")
parser.build_automaton()

print("\n[4/5] Building LR(0) tables...")
parser.build_lr0_tables()

print("\n[5/5] Building SLR(1) tables...")
parser.build_slr_tables()

print("\n" + "=" * 70)
print("RESULTS")
print("=" * 70)
print(f"States: {len(parser.states)}")
print(f"Transitions: {len(parser.transitions)}")
print(f"\nLR(0): {'✓ YES' if not parser.lr0_conflicts else '✗ NO'} ({len(parser.lr0_conflicts)} conflicts)")
print(f"SLR(1): {'✓ YES' if not parser.slr_conflicts else '✗ NO'} ({len(parser.slr_conflicts)} conflicts)")

print("\nExporting to Excel...")
parser.export_to_excel('lr0_slr_analysis.xlsx')

print("\n" + "=" * 70)
print("✓ COMPLETE!")
print("=" * 70)
