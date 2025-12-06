"""
CLR(1) and LALR(1) Parser Generator for Mini RecipeScript Grammar
Same grammar as LR(0)/SLR for comparison
Generates complete Excel with states, transitions, ACTION/GOTO tables
"""

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class LR1Item:
    """LR(1) Item: [A → α·β, a] with lookahead"""
    def __init__(self, lhs, rhs, dot, lookahead):
        self.lhs = lhs
        self.rhs = tuple(rhs)
        self.dot = dot
        self.lookahead = lookahead
    
    def __eq__(self, other):
        return (self.lhs == other.lhs and self.rhs == other.rhs and 
                self.dot == other.dot and self.lookahead == other.lookahead)
    
    def __hash__(self):
        return hash((self.lhs, self.rhs, self.dot, self.lookahead))
    
    def __repr__(self):
        rhs_list = list(self.rhs) if self.rhs else ['ε']
        rhs_with_dot = rhs_list[:self.dot] + ['•'] + rhs_list[self.dot:]
        return f"[{self.lhs} → {' '.join(rhs_with_dot)}, {self.lookahead}]"
    
    def core(self):
        """Return core (item without lookahead) for LALR"""
        return (self.lhs, self.rhs, self.dot)
    
    def next_symbol(self):
        if self.dot < len(self.rhs):
            return self.rhs[self.dot]
        return None
    
    def is_complete(self):
        return self.dot >= len(self.rhs)
    
    def advance(self):
        return LR1Item(self.lhs, self.rhs, self.dot + 1, self.lookahead)


class CLR_LALR_Parser:
    def __init__(self):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.start = None
        self.aug_start = None
        
        # CLR
        self.clr_states = []
        self.clr_state_map = {}
        self.clr_transitions = {}
        self.clr_action = {}
        self.clr_goto = {}
        self.clr_conflicts = []
        
        # LALR
        self.lalr_states = []
        self.lalr_state_map = {}
        self.lalr_transitions = {}
        self.lalr_action = {}
        self.lalr_goto = {}
        self.lalr_conflicts = []
        
        self.first = defaultdict(set)
    
    def add_production(self, lhs, rhs_list):
        if lhs not in self.grammar:
            self.grammar[lhs] = []
        self.non_terminals.add(lhs)
        for rhs in rhs_list:
            self.grammar[lhs].append(tuple(rhs))
    
    def is_terminal(self, sym):
        return sym not in self.non_terminals and sym != '$'
    
    def augment_grammar(self, start_symbol):
        self.start = start_symbol
        self.aug_start = start_symbol + "'"
        self.grammar[self.aug_start] = [(start_symbol,)]
        self.non_terminals.add(self.aug_start)
    
    def compute_first(self):
        """Compute FIRST sets"""
        for nt in self.non_terminals:
            self.first[nt] = set()
        
        changed = True
        while changed:
            changed = False
            for lhs, prods in self.grammar.items():
                for rhs in prods:
                    if not rhs:
                        if 'ε' not in self.first[lhs]:
                            self.first[lhs].add('ε')
                            changed = True
                    else:
                        for sym in rhs:
                            if self.is_terminal(sym):
                                if sym not in self.first[lhs]:
                                    self.first[lhs].add(sym)
                                    changed = True
                                break
                            else:
                                old_size = len(self.first[lhs])
                                self.first[lhs] |= (self.first[sym] - {'ε'})
                                if len(self.first[lhs]) > old_size:
                                    changed = True
                                if 'ε' not in self.first[sym]:
                                    break
    
    def first_of_string(self, symbols):
        """Compute FIRST of string"""
        result = set()
        for sym in symbols:
            if sym == '$' or self.is_terminal(sym):
                result.add(sym)
                return result
            result |= (self.first[sym] - {'ε'})
            if 'ε' not in self.first[sym]:
                return result
        result.add('ε')
        return result
    
    def closure(self, items):
        """Compute closure of LR(1) items"""
        closure_set = set(items)
        changed = True
        
        while changed:
            changed = False
            new_items = set()
            for item in closure_set:
                next_sym = item.next_symbol()
                if next_sym and next_sym in self.non_terminals:
                    beta_a = list(item.rhs[item.dot + 1:]) + [item.lookahead]
                    first_beta_a = self.first_of_string(beta_a)
                    
                    for prod in self.grammar.get(next_sym, []):
                        for lookahead in first_beta_a:
                            if lookahead != 'ε':
                                new_item = LR1Item(next_sym, prod, 0, lookahead)
                                if new_item not in closure_set:
                                    new_items.add(new_item)
                                    changed = True
            closure_set |= new_items
        
        return frozenset(closure_set)
    
    def goto(self, items, symbol):
        """Compute GOTO"""
        moved = set()
        for item in items:
            if item.next_symbol() == symbol:
                moved.add(item.advance())
        
        if moved:
            return self.closure(moved)
        return frozenset()
    
    def build_clr_automaton(self):
        """Build CLR(1) automaton"""
        print("\n=== Building CLR(1) Automaton ===")
        
        # Collect terminals
        for prods in self.grammar.values():
            for prod in prods:
                for sym in prod:
                    if self.is_terminal(sym):
                        self.terminals.add(sym)
        self.terminals.add('$')
        
        # Initial state
        initial_item = LR1Item(self.aug_start, self.grammar[self.aug_start][0], 0, '$')
        initial_state = self.closure({initial_item})
        
        self.clr_states = [initial_state]
        self.clr_state_map[initial_state] = 0
        queue = [initial_state]
        
        while queue:
            current = queue.pop(0)
            current_idx = self.clr_state_map[current]
            
            symbols = set()
            for item in current:
                sym = item.next_symbol()
                if sym:
                    symbols.add(sym)
            
            for sym in symbols:
                next_state = self.goto(current, sym)
                if next_state and next_state not in self.clr_state_map:
                    self.clr_state_map[next_state] = len(self.clr_states)
                    self.clr_states.append(next_state)
                    queue.append(next_state)
                
                if next_state:
                    self.clr_transitions[(current_idx, sym)] = self.clr_state_map[next_state]
        
        print(f"✓ CLR: Built {len(self.clr_states)} states")
        print(f"✓ CLR: Built {len(self.clr_transitions)} transitions")
    
    def build_lalr_automaton(self):
        """Build LALR(1) by merging CLR states with same cores"""
        print("\n=== Building LALR(1) Automaton ===")
        
        # Group CLR states by core
        core_groups = defaultdict(list)
        for idx, state in enumerate(self.clr_states):
            core = frozenset(item.core() for item in state)
            core_groups[core].append(idx)
        
        # Merge states with same core
        clr_to_lalr = {}
        lalr_idx = 0
        
        for core, clr_indices in core_groups.items():
            # Merge all items from states with same core
            merged_items = set()
            for clr_idx in clr_indices:
                merged_items.update(self.clr_states[clr_idx])
            
            lalr_state = frozenset(merged_items)
            self.lalr_states.append(lalr_state)
            
            for clr_idx in clr_indices:
                clr_to_lalr[clr_idx] = lalr_idx
            
            lalr_idx += 1
        
        # Map transitions
        for (from_clr, sym), to_clr in self.clr_transitions.items():
            from_lalr = clr_to_lalr[from_clr]
            to_lalr = clr_to_lalr[to_clr]
            self.lalr_transitions[(from_lalr, sym)] = to_lalr
        
        print(f"✓ LALR: Merged to {len(self.lalr_states)} states")
        print(f"✓ LALR: Built {len(self.lalr_transitions)} transitions")
    
    def get_production_number(self, lhs, rhs):
        """Get production number"""
        prod_num = 0
        for nt in sorted(self.grammar.keys()):
            for prod in self.grammar[nt]:
                if nt == lhs and prod == rhs:
                    return prod_num
                prod_num += 1
        return -1
    
    def build_clr_tables(self):
        """Build CLR(1) parsing tables"""
        print("\n=== Building CLR(1) Tables ===")
        
        for idx in range(len(self.clr_states)):
            self.clr_action[idx] = {t: [] for t in self.terminals}
            self.clr_goto[idx] = {nt: None for nt in self.non_terminals if nt != self.aug_start}
        
        for idx, state in enumerate(self.clr_states):
            for item in state:
                if item.is_complete():
                    if item.lhs == self.aug_start:
                        self.clr_action[idx]['$'].append(('accept',))
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        if item.lookahead in self.terminals:
                            self.clr_action[idx][item.lookahead].append(('reduce', prod_num))
                else:
                    sym = item.next_symbol()
                    if sym and self.is_terminal(sym):
                        if (idx, sym) in self.clr_transitions:
                            next_idx = self.clr_transitions[(idx, sym)]
                            shift_action = ('shift', next_idx)
                            if shift_action not in self.clr_action[idx][sym]:
                                self.clr_action[idx][sym].append(shift_action)
            
            for nt in self.non_terminals:
                if nt != self.aug_start and (idx, nt) in self.clr_transitions:
                    self.clr_goto[idx][nt] = self.clr_transitions[(idx, nt)]
        
        # Check conflicts
        for idx in range(len(self.clr_states)):
            for t in self.terminals:
                if len(self.clr_action[idx][t]) > 1:
                    self.clr_conflicts.append((idx, t, self.clr_action[idx][t]))
        
        if self.clr_conflicts:
            print(f"⚠ CLR: Found {len(self.clr_conflicts)} conflicts")
        else:
            print("✓ CLR: No conflicts!")
    
    def build_lalr_tables(self):
        """Build LALR(1) parsing tables"""
        print("\n=== Building LALR(1) Tables ===")
        
        for idx in range(len(self.lalr_states)):
            self.lalr_action[idx] = {t: [] for t in self.terminals}
            self.lalr_goto[idx] = {nt: None for nt in self.non_terminals if nt != self.aug_start}
        
        for idx, state in enumerate(self.lalr_states):
            for item in state:
                if item.is_complete():
                    if item.lhs == self.aug_start:
                        self.lalr_action[idx]['$'].append(('accept',))
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        if item.lookahead in self.terminals:
                            self.lalr_action[idx][item.lookahead].append(('reduce', prod_num))
                else:
                    sym = item.next_symbol()
                    if sym and self.is_terminal(sym):
                        if (idx, sym) in self.lalr_transitions:
                            next_idx = self.lalr_transitions[(idx, sym)]
                            shift_action = ('shift', next_idx)
                            if shift_action not in self.lalr_action[idx][sym]:
                                self.lalr_action[idx][sym].append(shift_action)
            
            for nt in self.non_terminals:
                if nt != self.aug_start and (idx, nt) in self.lalr_transitions:
                    self.lalr_goto[idx][nt] = self.lalr_transitions[(idx, nt)]
        
        # Check conflicts
        for idx in range(len(self.lalr_states)):
            for t in self.terminals:
                if len(self.lalr_action[idx][t]) > 1:
                    self.lalr_conflicts.append((idx, t, self.lalr_action[idx][t]))
        
        if self.lalr_conflicts:
            print(f"⚠ LALR: Found {len(self.lalr_conflicts)} conflicts")
        else:
            print("✓ LALR: No conflicts!")

    
    def export_to_excel(self, filename='clr_lalr_mini.xlsx'):
        """Export complete analysis to Excel"""
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        self.create_summary_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_clr_states_sheet(wb)
        self.create_clr_transitions_sheet(wb)
        self.create_clr_action_sheet(wb)
        self.create_clr_goto_sheet(wb)
        self.create_lalr_states_sheet(wb)
        self.create_lalr_transitions_sheet(wb)
        self.create_lalr_action_sheet(wb)
        self.create_lalr_goto_sheet(wb)
        
        wb.save(filename)
        print(f"\n✓ Excel file saved: {filename}")
    
    def create_summary_sheet(self, wb):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        ws['A1'] = "CLR(1) and LALR(1) Parser Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Grammar Statistics:"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A4'] = "Non-terminals:"
        ws['B4'] = len(self.non_terminals)
        ws['A5'] = "Terminals:"
        ws['B5'] = len(self.terminals)
        
        ws['A7'] = "CLR(1) Analysis:"
        ws['A7'].font = Font(bold=True, size=12, color="FF0000" if self.clr_conflicts else "008000")
        
        ws['A8'] = "States:"
        ws['B8'] = len(self.clr_states)
        ws['A9'] = "Transitions:"
        ws['B9'] = len(self.clr_transitions)
        ws['A10'] = "Is CLR(1)?"
        ws['B10'] = "NO ✗" if self.clr_conflicts else "YES ✓"
        ws['B10'].font = Font(bold=True, color="FF0000" if self.clr_conflicts else "008000")
        ws['A11'] = "Conflicts:"
        ws['B11'] = len(self.clr_conflicts)
        
        ws['A13'] = "LALR(1) Analysis:"
        ws['A13'].font = Font(bold=True, size=12, color="FF0000" if self.lalr_conflicts else "008000")
        
        ws['A14'] = "States:"
        ws['B14'] = len(self.lalr_states)
        ws['A15'] = "Transitions:"
        ws['B15'] = len(self.lalr_transitions)
        ws['A16'] = "Is LALR(1)?"
        ws['B16'] = "NO ✗" if self.lalr_conflicts else "YES ✓"
        ws['B16'].font = Font(bold=True, color="FF0000" if self.lalr_conflicts else "008000")
        ws['A17'] = "Conflicts:"
        ws['B17'] = len(self.lalr_conflicts)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Prod #"
        ws['B1'] = "LHS"
        ws['C1'] = "RHS"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="4472C4", fill_type="solid")
        
        row = 2
        prod_num = 0
        for lhs in sorted(self.grammar.keys()):
            for rhs in self.grammar[lhs]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = lhs
                ws[f'C{row}'] = ' '.join(rhs) if rhs else 'ε'
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
    
    def create_clr_states_sheet(self, wb):
        """Create CLR states sheet"""
        ws = wb.create_sheet("CLR States")
        
        ws['A1'] = "State"
        ws['B1'] = "LR(1) Items"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="5B9BD5", fill_type="solid")
        
        for idx, state in enumerate(self.clr_states):
            row = idx + 2
            ws[f'A{row}'] = f"I{idx}"
            ws[f'A{row}'].font = Font(bold=True)
            
            items_str = '\n'.join(str(item) for item in sorted(state, key=str))
            ws[f'B{row}'] = items_str
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = max(20, len(state) * 15)
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
    
    def create_clr_transitions_sheet(self, wb):
        """Create CLR transitions sheet"""
        ws = wb.create_sheet("CLR Transitions")
        
        ws['A1'] = "From"
        ws['B1'] = "Symbol"
        ws['C1'] = "To"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="ED7D31", fill_type="solid")
        
        row = 2
        for (from_st, sym), to_st in sorted(self.clr_transitions.items()):
            ws[f'A{row}'] = f"I{from_st}"
            ws[f'B{row}'] = sym
            ws[f'C{row}'] = f"I{to_st}"
            row += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
    
    def create_clr_action_sheet(self, wb):
        """Create CLR ACTION table"""
        ws = wb.create_sheet("CLR ACTION")
        
        terminals = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col, t in enumerate(terminals, 2):
            ws.cell(1, col, t)
        
        for col in range(1, len(terminals) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="5B9BD5", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.clr_states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, t in enumerate(terminals, 2):
                actions = self.clr_action[idx][t]
                if actions:
                    strs = []
                    for act in actions:
                        if act[0] == 'shift':
                            strs.append(f"s{act[1]}")
                        elif act[0] == 'reduce':
                            strs.append(f"r{act[1]}")
                        elif act[0] == 'accept':
                            strs.append("acc")
                    
                    cell = ws.cell(row, col, '\n'.join(strs))
                    cell.alignment = Alignment(wrap_text=True, horizontal='center')
                    
                    if len(actions) > 1:
                        cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(terminals) + 2):
            ws.column_dimensions[chr(64+col)].width = 12
    
    def create_clr_goto_sheet(self, wb):
        """Create CLR GOTO table"""
        ws = wb.create_sheet("CLR GOTO")
        
        nts = sorted([nt for nt in self.non_terminals if nt != self.aug_start])
        
        ws['A1'] = "State"
        for col, nt in enumerate(nts, 2):
            ws.cell(1, col, nt)
        
        for col in range(1, len(nts) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="5B9BD5", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.clr_states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, nt in enumerate(nts, 2):
                val = self.clr_goto[idx][nt]
                if val is not None:
                    cell = ws.cell(row, col, val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="D9EAD3", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(nts) + 2):
            ws.column_dimensions[chr(64+col)].width = 15
    
    def create_lalr_states_sheet(self, wb):
        """Create LALR states sheet"""
        ws = wb.create_sheet("LALR States")
        
        ws['A1'] = "State"
        ws['B1'] = "LR(1) Items (Merged)"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="70AD47", fill_type="solid")
        
        for idx, state in enumerate(self.lalr_states):
            row = idx + 2
            ws[f'A{row}'] = f"I{idx}"
            ws[f'A{row}'].font = Font(bold=True)
            
            items_str = '\n'.join(str(item) for item in sorted(state, key=str))
            ws[f'B{row}'] = items_str
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = max(20, len(state) * 15)
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
    
    def create_lalr_transitions_sheet(self, wb):
        """Create LALR transitions sheet"""
        ws = wb.create_sheet("LALR Transitions")
        
        ws['A1'] = "From"
        ws['B1'] = "Symbol"
        ws['C1'] = "To"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="FFC000", fill_type="solid")
        
        row = 2
        for (from_st, sym), to_st in sorted(self.lalr_transitions.items()):
            ws[f'A{row}'] = f"I{from_st}"
            ws[f'B{row}'] = sym
            ws[f'C{row}'] = f"I{to_st}"
            row += 1
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
    
    def create_lalr_action_sheet(self, wb):
        """Create LALR ACTION table"""
        ws = wb.create_sheet("LALR ACTION")
        
        terminals = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col, t in enumerate(terminals, 2):
            ws.cell(1, col, t)
        
        for col in range(1, len(terminals) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="70AD47", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.lalr_states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, t in enumerate(terminals, 2):
                actions = self.lalr_action[idx][t]
                if actions:
                    strs = []
                    for act in actions:
                        if act[0] == 'shift':
                            strs.append(f"s{act[1]}")
                        elif act[0] == 'reduce':
                            strs.append(f"r{act[1]}")
                        elif act[0] == 'accept':
                            strs.append("acc")
                    
                    cell = ws.cell(row, col, '\n'.join(strs))
                    cell.alignment = Alignment(wrap_text=True, horizontal='center')
                    
                    if len(actions) > 1:
                        cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(terminals) + 2):
            ws.column_dimensions[chr(64+col)].width = 12
    
    def create_lalr_goto_sheet(self, wb):
        """Create LALR GOTO table"""
        ws = wb.create_sheet("LALR GOTO")
        
        nts = sorted([nt for nt in self.non_terminals if nt != self.aug_start])
        
        ws['A1'] = "State"
        for col, nt in enumerate(nts, 2):
            ws.cell(1, col, nt)
        
        for col in range(1, len(nts) + 2):
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, col).fill = PatternFill(start_color="70AD47", fill_type="solid")
            ws.cell(1, col).alignment = Alignment(horizontal='center')
        
        for idx in range(len(self.lalr_states)):
            row = idx + 2
            ws.cell(row, 1, f"I{idx}").font = Font(bold=True)
            
            for col, nt in enumerate(nts, 2):
                val = self.lalr_goto[idx][nt]
                if val is not None:
                    cell = ws.cell(row, col, val)
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = PatternFill(start_color="D9EAD3", fill_type="solid")
        
        ws.column_dimensions['A'].width = 10
        for col in range(2, len(nts) + 2):
            ws.column_dimensions[chr(64+col)].width = 15


# Define mini grammar (same as LR0/SLR)
parser = CLR_LALR_Parser()

parser.add_production('S', [['D'], ['O']])
parser.add_production('D', [['ingredient', 'id', '=', 'E']])
parser.add_production('O', [['mix', 'id']])
parser.add_production('E', [['E', '+', 'T'], ['T']])
parser.add_production('T', [['num'], ['id']])

print("=" * 70)
print("CLR(1) AND LALR(1) PARSER ANALYSIS")
print("RecipeScript Mini Grammar")
print("=" * 70)

print("\nGrammar:")
for lhs in sorted(parser.grammar.keys()):
    for rhs in parser.grammar[lhs]:
        print(f"  {lhs} → {' '.join(rhs)}")

print("\n[1/6] Augmenting grammar...")
parser.augment_grammar('S')
print(f"✓ Augmented start: {parser.aug_start}")

print("\n[2/6] Computing FIRST sets...")
parser.compute_first()
print("✓ FIRST computed")

print("\n[3/6] Building CLR(1) automaton...")
parser.build_clr_automaton()

print("\n[4/6] Building LALR(1) automaton...")
parser.build_lalr_automaton()

print("\n[5/6] Building CLR(1) tables...")
parser.build_clr_tables()

print("\n[6/6] Building LALR(1) tables...")
parser.build_lalr_tables()

print("\n" + "=" * 70)
print("RESULTS")
print("=" * 70)
print(f"CLR(1):  {len(parser.clr_states)} states, {len(parser.clr_transitions)} transitions")
print(f"         {'✓ YES' if not parser.clr_conflicts else '✗ NO'} ({len(parser.clr_conflicts)} conflicts)")
print(f"\nLALR(1): {len(parser.lalr_states)} states, {len(parser.lalr_transitions)} transitions")
print(f"         {'✓ YES' if not parser.lalr_conflicts else '✗ NO'} ({len(parser.lalr_conflicts)} conflicts)")

print("\nExporting to Excel...")
parser.export_to_excel('clr_lalr_mini.xlsx')

print("\n" + "=" * 70)
print("✓ COMPLETE!")
print("=" * 70)
